"""
FinanceInternRadar 用户数据导出工具
将 SQLite 数据库中的订阅用户数据导出为格式化的 Excel 文件。
"""

import sqlite3
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "backend", "data", "radar.db")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "backend", "exports")
os.makedirs(OUTPUT_DIR, exist_ok=True)
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, f"用户数据_{TIMESTAMP}.xlsx")

# ─── Style definitions ───────────────────────────────
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1a1a2e")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=11, color="555555")


def style_sheet(ws, headers, col_widths, rows, zebra=True):
    """Apply consistent styling to a sheet."""
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Set column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if zebra and row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL

    # Freeze header row
    ws.freeze_panes = "A2"


def export():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    wb = Workbook()

    # ─── Sheet 1: Subscribers ─────────────────────────
    ws1 = wb.active
    ws1.title = "订阅用户"

    sub_headers = ["ID", "邮箱", "姓名", "注册时间", "订阅公司数", "订阅职位类型"]
    sub_widths = [6, 32, 16, 20, 14, 30]

    sub_rows = []
    sub_data = conn.execute("""
        SELECT id, email, name, created_at
        FROM subscribers
        ORDER BY created_at DESC
    """).fetchall()

    for s in sub_data:
        # Count subscriptions for this subscriber
        sub_count = conn.execute(
            "SELECT COUNT(*) FROM subscriptions WHERE subscriber_id = ?", (s["id"],)
        ).fetchone()[0]
        # Get subscribed role types
        role_types = conn.execute(
            "SELECT DISTINCT role_type FROM subscriptions WHERE subscriber_id = ? AND role_type IS NOT NULL",
            (s["id"],),
        ).fetchall()
        roles_str = ", ".join(r[0] for r in role_types) if role_types else "—"

        sub_rows.append([
            s["id"],
            s["email"],
            s["name"] or "—",
            s["created_at"],
            sub_count,
            roles_str,
        ])

    style_sheet(ws1, sub_headers, sub_widths, sub_rows)
    ws1.auto_filter.ref = ws1.dimensions

    # ─── Sheet 2: Subscriptions ───────────────────────
    ws2 = wb.create_sheet("订阅明细")

    sub_detail_headers = ["订阅ID", "用户邮箱", "用户姓名", "公司", "地区", "类别", "职位类型"]
    sub_detail_widths = [10, 32, 16, 28, 8, 18, 24]

    detail_rows = []
    details = conn.execute("""
        SELECT
            s.id AS sub_id,
            sub.email,
            sub.name AS sub_name,
            c.name AS company_name,
            c.region,
            c.category,
            s.role_type
        FROM subscriptions s
        JOIN subscribers sub ON s.subscriber_id = sub.id
        LEFT JOIN companies c ON s.company_id = c.id
        ORDER BY sub.email, c.name
    """).fetchall()

    for d in details:
        detail_rows.append([
            d["sub_id"],
            d["email"],
            d["sub_name"] or "—",
            d["company_name"] or f"公司ID:{d['sub_id']}",
            d["region"] or "—",
            d["category"] or "—",
            d["role_type"] or "全部",
        ])

    style_sheet(ws2, sub_detail_headers, sub_detail_widths, detail_rows)
    ws2.auto_filter.ref = ws2.dimensions

    # ─── Sheet 3: Summary ─────────────────────────────
    ws3 = wb.create_sheet("统计概览")

    # Stats
    total_subscribers = conn.execute("SELECT COUNT(*) FROM subscribers").fetchone()[0]
    total_subs = conn.execute("SELECT COUNT(*) FROM subscriptions").fetchone()[0]
    total_companies = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    today_new = conn.execute(
        "SELECT COUNT(*) FROM subscribers WHERE date(created_at) = date('now')"
    ).fetchone()[0]

    # By region
    region_stats = conn.execute("""
        SELECT c.region, COUNT(*) AS cnt
        FROM subscriptions s
        JOIN companies c ON s.company_id = c.id
        GROUP BY c.region
        ORDER BY cnt DESC
    """).fetchall()

    # By category
    cat_stats = conn.execute("""
        SELECT c.category, COUNT(*) AS cnt
        FROM subscriptions s
        JOIN companies c ON s.company_id = c.id
        GROUP BY c.category
        ORDER BY cnt DESC
    """).fetchall()

    # Top companies
    top_companies = conn.execute("""
        SELECT c.name, COUNT(*) AS cnt
        FROM subscriptions s
        JOIN companies c ON s.company_id = c.id
        GROUP BY c.name
        ORDER BY cnt DESC
        LIMIT 10
    """).fetchall()

    # Write summary header
    ws3.merge_cells("A1:D1")
    title_cell = ws3.cell(row=1, column=1, value="FinanceInternRadar 用户数据统计")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:D2")
    date_cell = ws3.cell(row=2, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.font = SUBTITLE_FONT
    date_cell.alignment = Alignment(horizontal="center")

    # Overview cards
    row = 4
    ws3.cell(row=row, column=1, value="核心指标").font = Font(name="微软雅黑", bold=True, size=12)
    ws3.merge_cells(f"A{row}:D{row}")
    row += 1

    metrics = [
        ("订阅用户总数", total_subscribers, "人"),
        ("订阅记录总数", total_subs, "条"),
        ("今日新增", today_new, "人"),
        ("监控公司数", total_companies, "家"),
    ]
    for i, (label, value, unit) in enumerate(metrics):
        col = i * 2 + 1  # A, C, E, G
        cell_label = ws3.cell(row=row, column=col, value=label)
        cell_label.font = Font(name="微软雅黑", size=10, color="666666")
        cell_label.alignment = CENTER_ALIGN

        cell_value = ws3.cell(row=row + 1, column=col, value=f"{value} {unit}")
        cell_value.font = Font(name="微软雅黑", bold=True, size=18, color="1a1a2e")
        cell_value.alignment = CENTER_ALIGN

        # Merge two rows for the value
        col_letter = get_column_letter(col)
        next_col = get_column_letter(col + 1)
        ws3.merge_cells(f"{col_letter}{row}:{next_col}{row}")
        ws3.merge_cells(f"{col_letter}{row + 1}:{next_col}{row + 1}")

    row += 3

    # Region breakdown
    ws3.cell(row=row, column=1, value="地区分布").font = Font(name="微软雅黑", bold=True, size=12)
    row += 1
    for col_idx, h in enumerate(["地区", "订阅数"], 1):
        cell = ws3.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 14
    row += 1
    for r in region_stats:
        ws3.cell(row=row, column=1, value=r["region"]).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=r["cnt"]).border = THIN_BORDER
        ws3.cell(row=row, column=2).alignment = CENTER_ALIGN
        row += 1

    row += 1

    # Category breakdown
    ws3.cell(row=row, column=1, value="类别分布").font = Font(name="微软雅黑", bold=True, size=12)
    row += 1
    for col_idx, h in enumerate(["类别", "订阅数"], 1):
        cell = ws3.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    row += 1
    for r in cat_stats:
        ws3.cell(row=row, column=1, value=r["category"]).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=r["cnt"]).border = THIN_BORDER
        ws3.cell(row=row, column=2).alignment = CENTER_ALIGN
        row += 1

    row += 1

    # Top companies
    ws3.cell(row=row, column=1, value="最受欢迎公司 TOP 10").font = Font(name="微软雅黑", bold=True, size=12)
    row += 1
    for col_idx, h in enumerate(["排名", "公司", "订阅数"], 1):
        cell = ws3.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws3.column_dimensions["C"].width = 28
    row += 1
    for rank, tc in enumerate(top_companies, 1):
        ws3.cell(row=row, column=1, value=rank).border = THIN_BORDER
        ws3.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws3.cell(row=row, column=2, value=tc["name"]).border = THIN_BORDER
        ws3.cell(row=row, column=3, value=tc["cnt"]).border = THIN_BORDER
        ws3.cell(row=row, column=3).alignment = CENTER_ALIGN
        row += 1

    conn.close()

    # Save
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = export()
    print(f"数据已导出到: {path}")